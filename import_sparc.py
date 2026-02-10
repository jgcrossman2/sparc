#!/usr/bin/env python3
"""Import SPARC.xlsx data into the sparc PostgreSQL database."""

import re
import openpyxl
import psycopg2

XLSX_PATH = "SPARC.xlsx"
DB_NAME = "sparc"

MEMBERSHIP_LEVELS = {
    "bush", "eisenhower", "lincoln", "reagan", "roosevelt",
    "family", "individual", "student",
}

# Patterns in attendance record that mean the member is inactive
INACTIVE_PATTERNS = re.compile(
    r"remove|moved|expired", re.IGNORECASE
)


def parse_attendance(raw):
    """Parse the Attendance Record column.

    Returns (membership_level, attendance_notes, active).
    """
    if raw is None:
        return None, None, True

    text = str(raw).strip()
    if not text:
        return None, None, True

    active = not bool(INACTIVE_PATTERNS.search(text))

    # Try to extract a membership level
    # Patterns like "member - Bush", "member - individual - Lives in DC"
    level = None
    notes = text

    match = re.match(r"member\s*-\s*(\w+)(.*)", text, re.IGNORECASE)
    if match:
        candidate = match.group(1).strip().lower()
        if candidate in MEMBERSHIP_LEVELS:
            level = candidate
            # Capitalise named levels, keep generic ones lowercase
            if candidate in ("bush", "eisenhower", "lincoln", "reagan", "roosevelt"):
                level = candidate.capitalize()
            remainder = match.group(2).strip()
            # Strip leading dash/hyphen from remainder
            remainder = re.sub(r"^-\s*", "", remainder).strip()
            notes = remainder if remainder else None
        elif candidate == "expired":
            # "member - expired" → no level, mark inactive
            remainder = match.group(2).strip()
            remainder = re.sub(r"^-\s*", "", remainder).strip()
            notes = text  # keep full text as notes
    elif text.lower() == "member":
        level = None
        notes = "member (level unspecified)"
    elif text.lower() == "student" or text.lower() == "high school student":
        level = "student"
        notes = text if text.lower() != "student" else None
    elif text.lower() == "stanford student":
        level = "student"
        notes = "Stanford Student"
    elif text.lower() == "new member":
        notes = text

    return level, notes, active


def clean_phone(raw):
    """Convert phone values (possibly float) to clean string."""
    if raw is None:
        return None
    if isinstance(raw, float):
        # e.g. 5103348246.0 → "5103348246"
        return str(int(raw))
    return str(raw).strip() or None


def clean_tickets(raw):
    """Convert ticket values to int or None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    # Handle backtick errors like "`"
    s = str(raw).strip()
    if s in ("`", "'`'", ""):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    conn = psycopg2.connect(dbname=DB_NAME)
    cur = conn.cursor()

    # Clear existing data
    cur.execute("DELETE FROM member;")

    insert_sql = """
        INSERT INTO member (
            firstname, lastname, phone, email, board_contact,
            membership_level, attendance_notes, outreach_notes,
            tickets_feb_2026, tickets_apr, active
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        firstname = row[0].value  # A: First name
        lastname = row[1].value   # B: Last name

        # Skip completely empty rows
        if not firstname and not lastname:
            continue

        firstname = str(firstname).strip() if firstname else None
        lastname = str(lastname).strip() if lastname else None

        phone = clean_phone(row[2].value)          # C: phone
        board_contact = str(row[3].value).strip() if row[3].value else None  # D: board member
        outreach_notes = str(row[4].value).strip() if row[4].value else None  # E: Outreach
        attendance_raw = row[5].value                # F: Attendance Record
        email = str(row[6].value).strip() if row[6].value else None  # G: Email

        tickets_feb = clean_tickets(row[7].value)    # H: Tickets (current Feb 2026 event)
        tickets_apr = clean_tickets(row[12].value)   # M: April tickets

        membership_level, attendance_notes, active = parse_attendance(attendance_raw)

        cur.execute(insert_sql, (
            firstname, lastname, phone, email, board_contact,
            membership_level, attendance_notes, outreach_notes,
            tickets_feb, tickets_apr, active,
        ))
        count += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"Imported {count} rows into member table.")


if __name__ == "__main__":
    main()
